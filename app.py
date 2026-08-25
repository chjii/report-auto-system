import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import re
import json
import os

st.set_page_config(page_title="자동창고 보고서 생성기", layout="centered")
st.title("📝 현장 보고서 자동 생성기")

# ==========================================
# 기억 장치 (현장 데이터 & 임시 저장 데이터)
# ==========================================
DATA_FILE = "site_memory.json"
DRAFT_FILE = "draft_memory.json"

def load_memory():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_memory(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_draft():
    if os.path.exists(DRAFT_FILE):
        with open(DRAFT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"contents": ""}

def save_draft(contents):
    with open(DRAFT_FILE, "w", encoding="utf-8") as f:
        json.dump({"contents": contents}, f, ensure_ascii=False)

memory_db = load_memory()

# ==========================================
# 0. 작업 분류 (공사 / 점검)
# ==========================================
st.markdown("### 📋 작업 분류 선택")
task_type = st.radio("어떤 종류의 보고서를 작성하시나요?", ["점검", "공사"], horizontal=True)

st.divider()

# ==========================================
# 1. 대분류 / 중분류 현장 선택
# ==========================================
st.markdown("### 🏢 업체 및 현장 정보")

site_dict = {
    "MXROBOTICS": ["태준제약"],
    "SFA SERVICE": ["BGF 로지스 광주 현장", "BGF 로지스 진천 현장"],
    "BLUEONE": ["유한킴벌리 충주공장"]
}

col_v, col_s = st.columns(2)
with col_v:
    vendor = st.selectbox("업체 대분류", list(site_dict.keys()))

saved_sites = [site for site, info in memory_db.items() if info.get("vendor") == vendor and site not in site_dict[vendor]]
current_site_list = site_dict[vendor] + saved_sites + ["직접 입력..."]

with col_s:
    site_select = st.selectbox("현장명 중분류", current_site_list)

default_address = ""
default_manager = "김주영 책임"

if site_select == "직접 입력...":
    site_name = st.text_input("새로운 현장명을 직접 입력해주세요 (필수)")
else:
    site_name = site_select
    if site_name in memory_db:
        default_address = memory_db[site_name].get("address", "")
        default_manager = memory_db[site_name].get("manager", "김주영 책임")

address = st.text_input("현장 주소", value=default_address, placeholder="예: 경기도 용인시...")

st.divider()

# ==========================================
# 2. 작업자 및 기간 입력 
# ==========================================
col1, col2 = st.columns(2)
with col1:
    author = st.text_input("작성자", value="지창현")
with col2:
    manager_list = ["김주영 책임", "최진명 차장", "조상길 부장"]
    if default_manager not in manager_list:
        manager_list.insert(0, default_manager)
    manager_list.append("직접 입력...")
    
    manager_select = st.selectbox("담당자 선택", manager_list, index=manager_list.index(default_manager))

if manager_select == "직접 입력...":
    manager = st.text_input("담당자명을 직접 입력해주세요 (필수)")
else:
    manager = manager_select

date_range = st.date_input("작업 일자 (기간 선택)", [])
date_str = ""
if len(date_range) == 2:
    start, end = date_range
    if start.month == end.month:
        date_str = f"{start.strftime('%y. %m. %d')} ~ {end.strftime('%d')}"
    else:
        date_str = f"{start.strftime('%y. %m. %d')} ~ {end.strftime('%m. %d')}"

workers = st.text_input("작업자명 및 인원", placeholder="예: 최진명 차장 외 6명")

st.divider()

# ==========================================
# 3. 설비 선택 및 기본 점검 항목 세팅
# ==========================================
st.markdown("### ⚙️ 점검 진행 설비 선택")
equipments = st.multiselect(
    "현장에서 점검한 설비를 모두 선택하세요", 
    ["STACKER CRANE", "CONVEYOR", "RGV", "LIFT"], 
    default=["STACKER CRANE"]
)

DEFAULT_TEXTS = {
    "STACKER CRANE": [
        "1. [STACKER CRANE] 점검 공통사항",
        "  1) 승강부,주행부,FORK부 구동 MOTOR 및 감속기 발열 상태 및 OIL 누유 상태 점검",
        "  2) CARRIAGE INNER ROLLER, GUIDE ROLLER 구름 상태 및 마모 상태 점검",
        "  3) FORK CHAIN TENSION 점검 및 C/F BEARING, MC GUIDE GREASE 도포",
        "  4) STC(기상반) 단자대 풀림 상태 CHECK 및 재조임",
        "  5) 주행부 구동 WHEEL,종동 WHEEL, GUIDE ROLLER 구름 상태 및 마모 상태 점검"
    ],
    "CONVEYOR": [
        "1. [CONVEYOR] 점검 공통사항",
        "  1) 구동 MOTOR 및 감속기 발열/소음 상태 및 OIL 누유 상태 점검",
        "  2) 체인/벨트 장력 상태 및 마모 상태 점검",
        "  3) 구동/종동 ROLLER 구름 상태 점검 및 베어링 소음 확인",
        "  4) 센서(광전, 근접 등) 취부 상태 및 동작 상태 점검"
    ],
    "RGV": [
        "1. [RGV] 점검 공통사항",
        "  1) 주행부 구동 MOTOR 발열 및 소음, 누유 상태 점검",
        "  2) 주행 WHEEL 및 GUIDE ROLLER 마모 상태 점검",
        "  3) 집전기(Collector) 마모 상태 및 단자대 조임 상태 점검",
        "  4) 충돌 방지 센서 및 통신 장치 상태 점검"
    ],
    "LIFT": [
        "1. [LIFT] 점검 공통사항",
        "  1) 승강 MOTOR 및 감속기 소음/발열, 누유 상태 점검",
        "  2) 승강 CHAIN 및 장력, 마모 상태 점검",
        "  3) GUIDE ROLLER 구름 상태 및 마모 상태 점검",
        "  4) 상/하한 리미트 센서 및 낙하 방지 장치 동작 상태 점검"
    ]
}

# ==========================================
# 4. 스마트 메모장 (자동 임시 저장)
# ==========================================
col_a, col_b = st.columns([7, 3])
with col_a:
    st.markdown(f"**{task_type} 내용 (설비별 자동 분류 및 2번부터 자동 넘버링)**")
with col_b:
    if st.button("🗑️ 내용 초기화"):
        st.session_state.contents = ""
        save_draft("")
        st.rerun()

if "contents" not in st.session_state:
    st.session_state.contents = load_draft().get("contents", "")

def update_draft():
    save_draft(st.session_state.contents)

contents = st.text_area(
    "S/C, CV, RGV, LIFT 등 키워드를 포함해서 적어주시면 코드가 알아서 각 설비 구역으로 나눠서 정리합니다.", 
    height=150,
    key="contents",
    on_change=update_draft
)

# ==========================================
# 5. 스마트 사진 업로드 (미리보기 UI)
# ==========================================
st.divider()
st.markdown("### 📷 현장 사진 업로드 (사진 대장용)")
st.caption("실제 엑셀 양식과 동일하게 2칸씩 나란히 배치됩니다. 1칸당 최대 2장의 사진을 넣을 수 있습니다.")

if "photo_blocks" not in st.session_state:
    st.session_state.photo_blocks = 4

photo_data = []

for i in range(0, st.session_state.photo_blocks, 2):
    cols = st.columns(2)
    for j in range(2):
        block_idx = i + j
        if block_idx < st.session_state.photo_blocks:
            with cols[j]:
                with st.container(border=True):
                    st.markdown(f"**[{block_idx+1}번 칸]**")
                    photos = st.file_uploader(f"➕ 사진 추가 (최대 2장)", type=['png', 'jpg', 'jpeg'], accept_multiple_files=True, key=f"photo_{block_idx}", label_visibility="collapsed")
                    
                    if photos:
                        p_cols = st.columns(2)
                        for p_idx, p_file in enumerate(photos[:2]):
                            with p_cols[p_idx]:
                                st.image(p_file, use_container_width=True)
                                p_file.seek(0)
                    
                    desc = st.text_area("설명", key=f"desc_{block_idx}", height=68, placeholder="이 칸의 설명을 입력하세요.", label_visibility="collapsed")
                    photo_data.append((block_idx, photos, desc))

st.write("")
if st.button("➕ 사진 칸 2개(1줄) 추가하기"):
    st.session_state.photo_blocks += 2
    st.rerun()

# 💡 [핵심] 병합 셀 에러를 완벽하게 차단하는 무적의 추적 함수!
def get_safe_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == 'MergedCell':
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return ws.cell(row=mr.min_row, column=mr.min_col)
    return cell

def sort_rules(text):
    is_need_replace = 1 if "교체 필요" in text else 0
    match = re.search(r'(#)?(\d+)호기', text)
    ho_number = int(match.group(2)) if match else 9999
    return (is_need_replace, ho_number)

def write_equipment_block(ws, defaults, user_lines, row_idx):
    block_size = len(defaults) + len(user_lines) + 1 
    local_row = (row_idx - 1) % 38 + 1
    
    if local_row + block_size > 37 and block_size <= 26:
        row_idx = ((row_idx - 1) // 38 + 1) * 38 + 12
    
    for i, df_text in enumerate(defaults):
        local_row = (row_idx - 1) % 38 + 1
        if local_row > 37: 
            row_idx = ((row_idx - 1) // 38 + 1) * 38 + 12
            
        cell = get_safe_cell(ws, row_idx, 2)
        cell.value = df_text
        cell.font = Font(name='맑은 고딕', size=11, bold=(i==0), color="000000")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row_idx += 1
        
    for idx, text in enumerate(user_lines):
        local_row = (row_idx - 1) % 38 + 1
        if local_row > 37: 
            row_idx = ((row_idx - 1) // 38 + 1) * 38 + 12
            
        cell = get_safe_cell(ws, row_idx, 2)
        cell.value = f"{idx + 2}. {text}"
        
        if "교체 필요" in text:
            cell.font = Font(name='맑은 고딕', size=11, bold=True, color="FF0000")
        elif "조치" in text or "교체" in text:
            cell.font = Font(name='맑은 고딕', size=11, bold=True, color="0000FF")
        else:
            cell.font = Font(name='맑은 고딕', size=11, bold=False, color="000000")
            
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row_idx += 1
    
    row_idx += 1 
    return row_idx

# ==========================================
# 6. 엑셀 생성 실행
# ==========================================
st.divider()
if st.button(f"🚀 {vendor} {task_type}보고서 생성하기", use_container_width=True):
    if not site_name or not date_str or not author or not manager:
        st.warning("작성자, 담당자, 현장명, 날짜는 필수입니다!")
    elif not equipments:
        st.warning("점검 진행 설비를 최소 1개 이상 선택해주세요!")
    else:
        with st.spinner("엑셀 파일을 만들고 있습니다..."):
            try:
                memory_db[site_name] = {"vendor": vendor, "address": address, "manager": manager}
                save_memory(memory_db)

                prefix_map = {"MXROBOTICS": "mxr", "SFA SERVICE": "sfa", "BLUEONE": "blueone"}
                template_filename = f"template_{prefix_map[vendor]}_{task_type}.xlsx"

                # --------------------------------------------------
                # [A] 보고서 엑셀 처리
                # --------------------------------------------------
                wb_report = openpyxl.load_workbook(template_filename)
                ws_report = wb_report.active

                # 💡 모든 입력에 get_safe_cell 적용!
                get_safe_cell(ws_report, 5, 3).value = site_name
                get_safe_cell(ws_report, 6, 3).value = address
                get_safe_cell(ws_report, 9, 3).value = f"{site_name} 정기 {task_type}" 
                get_safe_cell(ws_report, 10, 3).value = date_str
                get_safe_cell(ws_report, 10, 9).value = workers
                get_safe_cell(ws_report, 6, 8).value = author    
                get_safe_cell(ws_report, 8, 3).value = manager   
                
                for page in range(5):
                    for r in range(12, 38):
                        actual_r = r + (page * 38)
                        cell = get_safe_cell(ws_report, actual_r, 2)
                        cell.value = None
                        cell.font = Font(name='맑은 고딕', size=11, color="000000")
                
                raw_lines = [line.strip() for line in contents.split('\n') if line.strip()]
                sorted_lines = sorted(raw_lines, key=sort_rules)
                
                sc_lines, cv_lines, rgv_lines, lift_lines = [], [], [], []
                for line in sorted_lines:
                    u_line = line.upper()
                    if "RGV" in u_line: rgv_lines.append(line)
                    elif "LIFT" in u_line or "리프트" in u_line: lift_lines.append(line)
                    elif "CV" in u_line or "CONVEYOR" in u_line or "컨베이어" in u_line: cv_lines.append(line)
                    elif "S/C" in u_line or "STC" in u_line or "크레인" in u_line or "호기" in u_line: sc_lines.append(line)
                    else: sc_lines.append(line) 

                current_row = 12
                
                if "STACKER CRANE" in equipments:
                    current_row = write_equipment_block(ws_report, DEFAULT_TEXTS["STACKER CRANE"], sc_lines, current_row)
                if "CONVEYOR" in equipments:
                    current_row = write_equipment_block(ws_report, DEFAULT_TEXTS["CONVEYOR"], cv_lines, current_row)
                if "RGV" in equipments:
                    current_row = write_equipment_block(ws_report, DEFAULT_TEXTS["RGV"], rgv_lines, current_row)
                if "LIFT" in equipments:
                    current_row = write_equipment_block(ws_report, DEFAULT_TEXTS["LIFT"], lift_lines, current_row)

                output_report = io.BytesIO()
                wb_report.save(output_report)
                output_report.seek(0)
                
                # --------------------------------------------------
                # [B] 사진 대장 엑셀 처리
                # --------------------------------------------------
                output_photo = None
                
                photo_data.sort(key=lambda x: x[0]) 
                has_photo_data = any(len(photos) > 0 or desc.strip() for _, photos, desc in photo_data)
                
                if has_photo_data:
                    wb_photo = openpyxl.load_workbook('photo_template.xlsx')
                    ws_photo = wb_photo.active
                    
                    # 💡 사진 대장 정보 입력에도 안전 함수 적용
                    get_safe_cell(ws_photo, 5, 2).value = f"{site_name} 자동화 창고 {task_type} 사진"
                    get_safe_cell(ws_photo, 10, 9).value = f"1. 현장명 : {site_name}"
                    get_safe_cell(ws_photo, 14, 9).value = f"3. 작업일자 : {date_str}"
                    get_safe_cell(ws_photo, 15, 9).value = f"4. 작업인원 : {workers}"

                    PHOTO_PAGE_ROWS = 85 

                    for i, (_, photos, desc) in enumerate(photo_data):
                        if not photos and not desc:
                            continue
                            
                        page = i // 4
                        block_idx = i % 4
                        base_row = (page * PHOTO_PAGE_ROWS) + 10 + (block_idx * 19)
                        desc_row = base_row + 16
                        
                        if desc:
                            cell = get_safe_cell(ws_photo, desc_row, 2)
                            cell.value = desc
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            
                        for j, photo_file in enumerate(photos):
                            if j >= 2: break 
                            
                            photo_file.seek(0) 
                            col = 'B' if j == 0 else 'D'
                            
                            img_pil = PILImage.open(photo_file)
                            img_pil.thumbnail((350, 350)) 
                            img_byte_arr = io.BytesIO()
                            img_pil.save(img_byte_arr, format='PNG')
                            img_byte_arr.seek(0)
                            xl_img = Image(img_byte_arr)
                            ws_photo.add_image(xl_img, f"{col}{base_row}")
                            
                    output_photo = io.BytesIO()
                    wb_photo.save(output_photo)
                    output_photo.seek(0)

                # --------------------------------------------------
                # [C] 다운로드 버튼 출력
                # --------------------------------------------------
                st.success("🎉 생성이 완료되었습니다!")
                
                col1, col2 = st.columns(2)
                vendor_prefix = prefix_map[vendor].upper()
                
                with col1:
                    st.download_button(
                        label=f"📥 [{task_type} 보고서] 다운로드",
                        data=output_report,
                        file_name=f"({vendor_prefix}_{task_type}보고서){site_name}_{date_str[:8].replace('. ', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                if output_photo:
                    with col2:
                        st.download_button(
                            label=f"🖼️ [{task_type} 사진] 다운로드",
                            data=output_photo,
                            file_name=f"({vendor_prefix}_{task_type}사진){site_name}_{date_str[:8].replace('. ', '')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

            except Exception as e:
                st.error(f"에러 발생! 깃허브에 템플릿 파일이 모두 있는지 확인해주세요. 에러내용: {e}")
